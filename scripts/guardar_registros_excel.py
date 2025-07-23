import pandas as pd
import os
import re
import openpyxl
import sys
from copy import copy
from datetime import datetime
from filtrador import filtrar_datos
from dotenv import load_dotenv
from cloud_storage import subir_archivo, descargar_archivo, existe_archivo

def buscar_archivo(ruta, nombre_archivo):
    try:
        for archivo in os.listdir(ruta):
            if re.search(nombre_archivo,archivo):
                return 'Encontrado'
        return 'No encontrado'
    except FileNotFoundError as desc:
        print (desc)
        return 'Path invalido'

def existe_hoja(ruta, nombre_hoja):
    try:
        wb = openpyxl.load_workbook(ruta)
        hoja_wb = wb[nombre_hoja]
        return True
    except KeyError as desc:
        print (desc)
        return False

def copiar_formato(almacenamiento, ruta, nombre_hoja):
    wb_nuevo = openpyxl.load_workbook(ruta)
    hoja_nueva = wb_nuevo[nombre_hoja]


    if (almacenamiento == "nube"):
        descargar_archivo("../plantilla/plantilla_tours.xlsx", "plantilla/plantilla_tours.xlsx", "tours-automaticos")

    wb_plantilla = openpyxl.load_workbook('../plantilla/plantilla_tours.xlsx')
    hoja_plantilla = wb_plantilla['Plantilla']

    for columna in hoja_plantilla.column_dimensions:
        hoja_nueva.column_dimensions[columna].width = hoja_plantilla.column_dimensions[columna].width
        
        for fila in hoja_plantilla.iter_rows():
            for celda in fila:
                nueva_celda = hoja_nueva.cell(row=celda.row, column=celda.column, value=celda.value)
                if celda.has_style:
                    nueva_celda.font = copy(celda.font)
                    nueva_celda.border = copy(celda.border)
                    nueva_celda.fill = copy(celda.fill)
                    nueva_celda.number_format = copy(celda.number_format)
                    nueva_celda.protection = copy(celda.protection)
                    nueva_celda.alignment = copy(celda.alignment)

    wb_nuevo.save(ruta)

def copiar_datos(df, ruta, nombre_hoja, fila_inicial):
    wb = openpyxl.load_workbook(ruta)
    hoja_wb = wb[nombre_hoja]

    for i, fila in enumerate(df.itertuples(index=False), start=fila_inicial):
        hoja_wb.insert_rows(idx=i, amount=1)
        for j, valor in enumerate(fila, start=1):
            hoja_wb.cell(row=i, column=j, value=valor)
    
    wb.save(ruta)

def guardar_registros(campus, df, fecha_solicitada, almacenamiento, descargado):
    meses = {
    "01": 'enero', "02": 'febrero', "03": 'marzo', "04": 'abril',
    "05": 'mayo', "06": 'junio', "07": 'julio', "08": 'agosto',
    "09": 'septiembre', "10": 'octubre', "11": 'noviembre', "12": 'diciembre'
    }
    
    dia = fecha_solicitada.strftime("%d")
    mes = meses[fecha_solicitada.strftime("%m")]
    año = fecha_solicitada.strftime("%Y")
    nombre_hoja = dia + " de " + mes
    nombre_excel = "Tour puertas abiertas " + mes + " " + año
    ruta_excel = '../data/' + nombre_excel + '.xlsx'
    fila_inicial = 8 if campus == "sur" else 4

    if (almacenamiento == "local"):
        busqueda = buscar_archivo("../data", nombre_excel)
    elif (almacenamiento == "nube"):
        if descargado:
            busqueda = buscar_archivo("../data", nombre_excel)
        else:
            busqueda = existe_archivo(f"data/{nombre_excel}.xlsx", "tours-automaticos")

    if busqueda == 'Encontrado':
        if (almacenamiento == "nube" and descargado == False):
            descargar_archivo(ruta_excel, f"data/{nombre_excel}.xlsx", "tours-automaticos")
        if existe_hoja(ruta_excel, nombre_hoja):
            print('Existe la hoja')
            copiar_datos(df, ruta_excel, nombre_hoja, fila_inicial)
            return nombre_excel
        else:
            print('No existe la hoja')
            wb = openpyxl.load_workbook(ruta_excel)
            wb.create_sheet(nombre_hoja)
            wb.save(ruta_excel)
    elif busqueda == 'No encontrado' :
        wb = openpyxl.Workbook()
        hoja_nueva = wb.active
        hoja_nueva.title = nombre_hoja
        wb.save(ruta_excel)
    elif busqueda == 'Path invalido':
        sys.exit(1)
    
    copiar_formato(almacenamiento, ruta_excel, nombre_hoja)
    copiar_datos(df, ruta_excel, nombre_hoja, fila_inicial)

    return nombre_excel

def registros_excel():
    existen_registros_norte = False
    existen_registros_sur = False
    descargado = False
    nombre_excel = None
    load_dotenv()
    almacenamiento = os.getenv("MODO_ALMACENAMIENTO")
    
    print("inicia filtrado de datos") 
    df_norte, df_sur = filtrar_datos("../data/registros_tours.csv")
    print("termina filtrado de datos")
    fecha_solicitada = datetime(2025, 5, 24).date()

    df_sur_fecha = df_sur[df_sur['Día de visita Sur_standar'].dt.date == fecha_solicitada]
    if not df_sur_fecha.empty:
        df_sur_fecha = df_sur_fecha.drop('Día de visita Sur_standar', axis=1)
        df_sur_fecha = df_sur_fecha.sort_values(by='Nombre')
        print("se guardan registros sur")
        nombre_excel = guardar_registros("sur", df_sur_fecha, fecha_solicitada, almacenamiento, descargado)
        print("se guardaron los registros sur")
        existen_registros_sur = True
        descargado = True

    df_norte_fecha = df_norte[df_norte['Día de visita Norte_standar'].dt.date == fecha_solicitada]
    if not df_norte_fecha.empty:
        df_norte_fecha = df_norte_fecha.drop('Día de visita Norte_standar', axis=1)
        df_norte_fecha = df_norte_fecha.sort_values(by='Nombre')
        print("se guardan registros norte")
        if not nombre_excel: 
            nombre_excel = guardar_registros("norte", df_norte_fecha, fecha_solicitada, almacenamiento, descargado)
        else: 
            guardar_registros("norte", df_norte_fecha, fecha_solicitada, almacenamiento, descargado)
        print("se guardaron registros norte")
        existen_registros_norte = True

    if (almacenamiento == "nube" and (existen_registros_norte or existen_registros_sur)):
        subir_archivo(f"../data/{nombre_excel}.xlsx", f"data/{nombre_excel}.xlsx", "tours-automaticos")
        nombre_excel = nombre_excel + ".xlsx"
    fecha_solicitada_string = fecha_solicitada.strftime("%d %m %Y")
    print("todo listo en Excel")
    return fecha_solicitada_string, existen_registros_norte, existen_registros_sur, nombre_excel
